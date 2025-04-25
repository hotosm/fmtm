# Copyright (c) Humanitarian OpenStreetMap Team
#
# This file is part of osm_fieldwork.
#
#     osm-fieldwork is free software: you can redistribute it and/or modify
#     it under the terms of the GNU General Public License as published by
#     the Free Software Foundation, either version 3 of the License, or
#     (at your option) any later version.
#
#     osm-fieldwork is distributed in the hope that it will be useful,
#     but WITHOUT ANY WARRANTY; without even the implied warranty of
#     MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#     GNU General Public License for more details.
#
#     You should have received a copy of the GNU General Public License
#     along with osm_fieldwork.  If not, see <https:#www.gnu.org/licenses/>.
#
"""Test functionality of update_form.py."""

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook, worksheet
from pyxform.xls2xform import convert as xform_convert

from osm_fieldwork.update_xlsform import append_field_mapping_fields
from osm_fieldwork.xlsforms import buildings, healthcare
from osm_fieldwork.form_components.translations import INCLUDED_LANGUAGES


async def test_merge_mandatory_fields():
    """Merge the mandatory fields XLSForm to a test survey form."""
    test_form = Path(__file__).parent / "test_data" / "test_form_for_mandatory_fields.xls"

    with open(test_form, "rb") as xlsform:
        form_bytes = BytesIO(xlsform.read())

    xformid, updated_form = await append_field_mapping_fields(form_bytes, "buildings")
    workbook = load_workbook(filename=BytesIO(updated_form.getvalue()))
    # Write merged xlsform to file for debugging
    with open("merged_xlsform.xlsx", "wb") as merged_xlsform:
        merged_xlsform.write(updated_form.getvalue())

    check_survey_sheet(workbook)
    # NOTE the choices sheet can have duplicates in the 'name' field without issue
    check_entities_sheet(workbook)
    check_form_title(workbook)

    # Check it's still a valid xlsform by converting to XML
    xform_convert(updated_form)
    check_translation_fields(workbook)


async def test_add_extra_select_from_file():
    """Append extra select_one_from_file questions based on Entity list names."""
    test_form = Path(__file__).parent / "test_data" / "test_form_for_mandatory_fields.xls"

    with open(test_form, "rb") as xlsform:
        form_bytes = BytesIO(xlsform.read())

    xformid, updated_form = await append_field_mapping_fields(form_bytes, "buildings", additional_entities=["roads", "waterpoints"])
    workbook = load_workbook(filename=BytesIO(updated_form.getvalue()))

    survey_sheet = workbook["survey"]
    name_column = [cell.value for cell in survey_sheet["B"]]

    assert "roads" in name_column, "The 'roads' field was not added to the survey sheet."
    assert "waterpoints" in name_column, "The 'waterpoints' field was not added to the survey sheet."


async def test_buildings_xlsform():
    """Merge and test if buildings form is a valid XLSForm."""
    with open(buildings, "rb") as xlsform:
        form_bytes = BytesIO(xlsform.read())
    xformid, updated_form = await append_field_mapping_fields(form_bytes, "buildings")
    # Check it's still a valid xlsform by converting to XML
    xform_convert(updated_form)

    workbook = load_workbook(filename=BytesIO(updated_form.getvalue()))
    check_translation_fields(workbook)


async def test_healthcare_xlsform():
    """Merge and test if buildings form is a valid XLSForm."""
    with open(healthcare, "rb") as xlsform:
        form_bytes = BytesIO(xlsform.read())
    xformid, updated_form = await append_field_mapping_fields(form_bytes, "healthcare")
    # Check it's still a valid xlsform by converting to XML
    xform_convert(updated_form)


def check_survey_sheet(workbook: Workbook) -> None:
    """Check the 'survey' sheet values and ensure no duplicates in 'name' column."""
    survey_sheet = get_sheet(workbook, "survey")
    name_col_index = get_column_index(survey_sheet, "name")
    calculation_col_index = get_column_index(survey_sheet, "calculation")
    check_for_duplicates(survey_sheet, name_col_index)


def check_entities_sheet(workbook: Workbook) -> None:
    """Check the 'entities' sheet values."""
    entities_sheet = get_sheet(workbook, "entities")
    label_col_index = get_column_index(entities_sheet, "label")

    test_label_present = any(
        row[0].value == "test label"
        for row in entities_sheet.iter_rows(min_col=label_col_index, max_col=label_col_index, min_row=2)
    )
    assert not test_label_present, "'test label' found in the 'label' column of 'entities' sheet."


def check_form_title(workbook: Workbook) -> None:
    """Check if the form_title is set correctly in the 'settings' sheet."""
    settings_sheet = get_sheet(workbook, "settings")
    form_title_col_index = get_column_index(settings_sheet, "form_title")

    form_title_value = settings_sheet.cell(row=2, column=form_title_col_index).value
    # NOTE previously we would strip 's' from plurals, but we no longer do this
    assert form_title_value == "buildings", "form_title field is not set to 'building'"


def check_translation_fields(workbook: Workbook):
    """Check if translation fields for all included languages were correctly matched."""
    survey_sheet = workbook["survey"]
    translation_found = {lang: False for lang in list(INCLUDED_LANGUAGES.keys())}

    # Iterate through the survey sheet columns and rows
    for row in survey_sheet.iter_rows(min_row=1, max_col=survey_sheet.max_column):
        for cell in row:
            # Check if the label field matches a translation field for each language
            for lang, code in INCLUDED_LANGUAGES.items():
                lang_label = f"label::{lang}({code})"
                if cell.value == lang_label:
                    translation_found[lang] = True

            # Ensure that the base 'label' field is no longer present
            if cell.value == "label":
                assert False, "The label field should be replaced by translated fields"

    # Check that all translations for the languages are present
    missing_translations = [lang for lang, found in translation_found.items() if not found]
    assert not missing_translations


def get_sheet(workbook: Workbook, sheet_name: str) -> worksheet.worksheet.Worksheet:
    """Helper function to get a sheet or raise an error."""
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"The '{sheet_name}' sheet was not found in the workbook")
    return workbook[sheet_name]


def check_for_duplicates(sheet: worksheet.worksheet.Worksheet, col_index: int) -> None:
    """Check for any duplicate values in a specific column of a sheet, ignoring None values."""
    seen_values = set()
    for row in sheet.iter_rows(min_col=col_index, max_col=col_index, min_row=2):
        value = row[0].value
        if value is None:
            # Skip None values, allowing them to appear multiple times
            continue
        if value in seen_values:
            raise AssertionError(f"Duplicate value '{value}' found in column '{col_index}' of sheet '{sheet.title}'.")
        seen_values.add(value)


def get_column_index(sheet: worksheet.worksheet.Worksheet, column_name: str) -> int:
    """Get the column index for the given column name."""
    for col_idx, col in enumerate(sheet.iter_cols(1, sheet.max_column), start=1):
        if col[0].value == column_name:
            return col_idx
    raise ValueError(f"Column '{column_name}' not found.")


def get_row_index(sheet: worksheet.worksheet.Worksheet, column_index: int, value: str) -> int:
    """Get the row index where the given column has the specified value."""
    for row_idx, row in enumerate(sheet.iter_rows(min_col=column_index, max_col=column_index), start=1):
        if row[0].value == value:
            return row_idx
    raise ValueError(f"Value '{value}' not found in column {column_index}.")
