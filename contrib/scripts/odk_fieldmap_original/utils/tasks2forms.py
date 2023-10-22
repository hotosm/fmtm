# Copyright (c) 2022, 2023 Humanitarian OpenStreetMap Team
# This file is part of FMTM.
#
#     FMTM is free software: you can redistribute it and/or modify
#     it under the terms of the GNU General Public License as published by
#     the Free Software Foundation, either version 3 of the License, or
#     (at your option) any later version.
#
#     FMTM is distributed in the hope that it will be useful,
#     but WITHOUT ANY WARRANTY; without even the implied warranty of
#     MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#     GNU General Public License for more details.
#
#     You should have received a copy of the GNU General Public License
#     along with FMTM.  If not, see <https:#www.gnu.org/licenses/>.
#

#!/bin/python3

"""If run from CLI, takes two positional arguments:

1) An input directory containing a subdirectory
full of individual GeoJSON files representing tasks
2) A ODK-compatible xlsform template

It then creates a subdirectory called 'forms'
and populates it with xlsforms specifically
referencing the individual GeoJSON files.
"""

import os
import sys

from openpyxl import load_workbook


def task_areas_to_forms(indir, form_template):
    """Accepts a project directory, with a subdirectory
    called /geojson full of GeoJSON point files,
    each representing an individual task (e.g. a batch
    of buildings or amenities to be visited for data
    collection) and returns a set of xlsforms, one for
    each task, with the appropriate references in
    place for the Select from Map question in ODK.
    """
    geojsondir = os.path.join(indir, "geojson")
    print(geojsondir)
    if not os.path.exists(geojsondir):
        print("Can't find geojson subdirectory")
        # Maybe throw an exception
    filelist = os.listdir(geojsondir)
    # keep only the files that are GeoJSON
    # TODO: Maybe reject all non-point input files;
    # kind of a hassle involving a spatial lib
    # but perhaps worth it
    aois = [x for x in filelist if os.path.splitext(x)[1].lower() == ".geojson"]
    outdir = os.path.join(indir, "forms")
    if not os.path.exists(outdir):
        print(f"Making directory {outdir}")
        os.makedirs(outdir)
    for aoi in aois:
        prep_form(form_template, aoi, outdir)


def prep_form(form_template, AOIfile, outdir):
    """Creates a modified copy of an ODK xlsform to refer
    to a specific area and GeoJSON file of features.
    Only works with forms built with this script in mind.
    """
    form_basename = os.path.splitext(os.path.basename(form_template))[0]
    print(form_basename)
    (AOIpath, AOIext) = os.path.splitext(AOIfile)
    AOIbasename = os.path.splitext(os.path.basename(AOIfile))[0]
    wb = load_workbook(filename=form_template)
    surveyws = wb["survey"]
    settingws = wb["settings"]

    # Make the form title the task name
    settingws["A2"] = f"{AOIbasename}"
    settingws["B2"] = f"{AOIbasename}"
    settingws["C2"] = f"{AOIbasename}"

    # Replace path to root of GeoJSON with the appropriate filename
    for row in surveyws.iter_rows():
        for cell in row:
            s = cell.value
            if s is not None:
                # check for cells referencing the geojson file
                if "instance('buildings" in s:
                    cell.value = s.replace("buildings", f"{AOIbasename}")
                if "select_one_from_file" in s:
                    cell.value = "select_one_from_file " + f"{AOIbasename}{AOIext}"

    # Write the individual XLSForm
    outfile = os.path.join(outdir, f"{AOIbasename}.xlsx")
    print(f"Writing: {outfile}")
    wb.save(outfile)


if __name__ == "__main__":
    """ """
    indir = sys.argv[1]
    formfile = sys.argv[2]
    print("here goes")
    task_areas_to_forms(indir, formfile)
